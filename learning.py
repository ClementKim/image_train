import os
import torch
import numpy as np

from torch import nn
from torch.utils.data import Dataset, DataLoader

from os import listdir

from openpyxl import load_workbook

lst = []

class CustumDataset(Dataset):
    def __init__(self, image, label, device, transform=None, target_tranform=None):
        self.image = image
        self.label = label
        self.device = device
        self.transform = transform
        self.target_transform = target_tranform

    def __len__(self) -> int:
        return len(self.label)

    def __getitem__(self, idx: int) -> tuple:
        image = torch.from_numpy(self.image[idx]).to(self.device).float()
        label = torch.tensor(self.label[idx], dtype=torch.long).to(self.device)
        return image, label


class NeuralNetwork(nn.Module):
    def __init__(self):
        super(NeuralNetwork, self).__init__()
        self.flatten = nn.Flatten()
        self.linear_relu_stack = nn.Sequential(
            nn.Linear(255 * 255 * 4, 255),
            nn.ReLU(),
            nn.Linear(255, 255),
            nn.ReLU(),
            nn.Linear(255, 3),
        )

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x = self.flatten(x)
        logits = self.linear_relu_stack(x)
        return logits


def train_loop(dataloader, model, loss_fn, optimizer):
    size = len(dataloader.dataset)

    for batch, (X, y) in enumerate(dataloader):
        pred = model(X)
        loss = loss_fn(pred, y)

        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

        if batch % 100 == 0:
            loss, current = loss.item(), (batch + 1) * len(X)
            print(f"loss: {loss:>7f} [{current:>5d}/{size:>5d}]")


def test_loop(dataloader, model, loss_fn):
    size = len(dataloader.dataset)
    num_batches = len(dataloader)
    test_loss, correct = 0, 0

    with torch.no_grad():
        for X, y in dataloader:
            pred = model(X)
            test_loss += loss_fn(pred, y).item()
            correct += torch.eq(pred.argmax(1), y).float().sum().item()

    test_loss /= num_batches
    correct /= size

    print(f"Test Error: \n Accuracy: {(100 * correct):>0.2f}%, Avg loss: {test_loss:.2f}\n")
    
    lst.append(100*correct)

def loading_npy(home_dir, optimizer_ipt, learning_rate, loss_fn_ipt, epochs, file_number):
    dataset_dir = os.path.join(home_dir, "dataset")
    required_file = ["train_image.npy", "train_labels.npy", "test_image.npy", "test_labels.npy"]
    if not ("dataset" in listdir(home_dir)) and all(file in listdir(dataset_dir) for file in required_file):
        print("dataset doesn't exist")
        exit()

    else:
        train_images = np.load(os.path.join(dataset_dir, 'train_image.npy'))
        print("train_image.npy loaded")

        train_labels = np.load(os.path.join(dataset_dir, 'train_labels.npy'))
        print("train_labels.npy loaded")

        test_images = np.load(os.path.join(dataset_dir, 'test_image.npy'))
        print("test_image.npy loaded")

        test_labels = np.load(os.path.join(dataset_dir, 'test_labels.npy'))
        print("test_labels.npy loaded")

        learning(home_dir, train_images, train_labels, test_images, test_labels, optimizer_ipt, learning_rate, loss_fn_ipt, epochs, file_number)


def learning(home_dir, train_images, train_labels, test_images, test_labels, optimizer_ipt, learning_rate, loss_fn_ipt, epochs, file_number):
    label_map = {
        0: "red light",
        1: "yellow light",
        2: "green light"
    }

    device = torch.device(
        "cuda"
        if torch.cuda.is_available()
        # else "mps"
        # if torch.backends.mps.is_available()
        else "cpu"
    )

    print(f"using {device}")

    train_data = CustumDataset(train_images, train_labels, device)
    test_data = CustumDataset(test_images, test_labels, device)

    check = [len(train_data.image) == len(train_images),
             len(train_data.label) == len(train_labels),
             len(test_data.image) == len(test_images),
             len(test_data.label) == len(test_labels)]

    if not all(check):
        print(f"{check.count(False)} errors")
        error = [i for i in range(len(check)) if not (check[i])]
        print(f"error index: {error}")

        exit()

    train_dataloader = DataLoader(train_data, batch_size=32, shuffle=True)
    test_dataloader = DataLoader(test_data, batch_size=32, shuffle=True)

    model = NeuralNetwork()
    model.to(device)
    print(model)

    model_dir = os.path.join(home_dir, "model")
    if not ("model" in listdir(home_dir)):
        os.mkdir(model_dir)

    optimizer = torch.optim.SGD(model.parameters(), lr=float(learning_rate))

    loss_fn = nn.CrossEntropyLoss().to(device)

    model.train()
    for t in range(int(epochs)):
        print(f"Epoch {t + 1}\n----------------")

        train_loop(train_dataloader, model, loss_fn, optimizer)
        test_loop(test_dataloader, model, loss_fn)

    print("learning is done\n")

    model_save_path = os.path.join(model_dir, str(loss_fn)[:-2] +
                                   "_" + str(epochs) + "_" + file_number + ".pt")

    torch.save(model.state_dict(), model_save_path)

    torch.cuda.empty_cache()
        
    rst_dir = os.path.join(os.getcwd(), "result")
    wb = load_workbook(os.path.join(rst_dir, "result.xlsx"), read_only = False, data_only = False)
    ws = wb.active
    
    ws["A" + str(int(file_number) + 1)] = loss_fn_ipt
    ws["B" + str(int(file_number) + 1)] = optimizer_ipt
    ws["C" + str(int(file_number) + 1)] = learning_rate
    ws["D" + str(int(file_number) + 1)] = epochs
    ws["E" + str(int(file_number) + 1)] = sum(lst)/len(lst)
    
    wb.save(os.path.join(rst_dir, "result.xlsx"))

